#!/usr/bin/env python3
"""
Qt-based Evaluation GUI for Excel files with translation data.

Reads Excel files where:
- Row 1: Headers (Source, Model 1, Model 2)
- Row 2+: Data rows with source text and translations

Features:
- Randomized display of translations (A/B) to prevent bias
- Auto-save on Next when both scores are non-zero
- LCS-based word matching to highlight common words between translations
"""
import sys
import json
import os
import random
from pathlib import Path
from collections import defaultdict

try:
    from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                                 QHBoxLayout, QLabel, QPushButton, QComboBox, 
                                 QSpinBox, QTextEdit, QRadioButton, QButtonGroup,
                                 QFileDialog, QMessageBox, QScrollArea, QGroupBox)
    from PyQt5.QtCore import Qt, QTimer
    from PyQt5.QtGui import QFont, QTextCharFormat, QColor, QTextCursor
except ImportError:
    print("ERROR: PyQt5 is required.")
    print("Install with: python -m pip install PyQt5")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required.")
    print("Install with: python -m pip install openpyxl")
    sys.exit(1)


# Model names for the 2 translation columns
MODELS = [
    "Model 1",
    "Model 2"
]

# JSON output files
EVALUATED_ROWS_FILE = "evaluated_rows.json"
MODEL_SCORES_FILE = "model_scores.json"
DETAILED_COMMENTS_FILE = "detailed_comments.json"


def find_lcs_words(text1, text2):
    """
    Find the Longest Common Subsequence of words between two texts.
    Uses dynamic programming to find matching word sequences.
    
    Returns:
        tuple: (matched_indices_text1, matched_indices_text2)
               Lists of word indices that are part of the LCS
    """
    # Split texts into words (handle Arabic text properly)
    words1 = text1.split()
    words2 = text2.split()
    
    if not words1 or not words2:
        return set(), set()
    
    m, n = len(words1), len(words2)
    
    # Build LCS table
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if words1[i-1] == words2[j-1]:
                dp[i][j] = dp[i-1][j-1] + 1
            else:
                dp[i][j] = max(dp[i-1][j], dp[i][j-1])
    
    # Backtrack to find matched indices
    matched_idx1 = set()
    matched_idx2 = set()
    
    i, j = m, n
    while i > 0 and j > 0:
        if words1[i-1] == words2[j-1]:
            matched_idx1.add(i-1)
            matched_idx2.add(j-1)
            i -= 1
            j -= 1
        elif dp[i-1][j] > dp[i][j-1]:
            i -= 1
        else:
            j -= 1
    
    return matched_idx1, matched_idx2


def get_word_positions(text):
    """
    Get the start and end positions of each word in the text.
    Returns list of (start, end) tuples for each word.
    """
    positions = []
    words = text.split()
    current_pos = 0
    
    for word in words:
        # Find the word in the remaining text
        start = text.find(word, current_pos)
        if start != -1:
            end = start + len(word)
            positions.append((start, end))
            current_pos = end
    
    return positions


class TranslationEvaluatorQt(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Translation Evaluation - Score and Save")
        self.setGeometry(100, 100, 1400, 900)
        
        self.wb = None
        self.current_sheet = None
        self.current_row_idx = None
        self.max_content_rows = 0
        self.filepath = None
        
        self.model_widgets = []
        
        # Track the randomized display order: maps display index -> actual model index
        # e.g., [1, 0] means display slot 0 shows Model 2, slot 1 shows Model 1
        self.display_order = [0, 1]
        
        # Store original translations for current row (before shuffling)
        self.current_translations = []
        
        self.init_ui()
    
    def init_ui(self):
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        
        # Top controls
        top_layout = QHBoxLayout()
        
        load_btn = QPushButton("📁 Load Excel...")
        load_btn.setFont(QFont("Segoe UI", 11, QFont.Bold))
        load_btn.clicked.connect(self.load_excel)
        load_btn.setMinimumHeight(40)
        top_layout.addWidget(load_btn)
        
        top_layout.addSpacing(20)
        
        sheet_label = QLabel("Sheet:")
        sheet_label.setFont(QFont("Segoe UI", 11, QFont.Bold))
        top_layout.addWidget(sheet_label)
        
        self.sheet_combo = QComboBox()
        self.sheet_combo.setFont(QFont("Segoe UI", 11))
        self.sheet_combo.setMinimumWidth(250)
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)
        top_layout.addWidget(self.sheet_combo)
        
        top_layout.addSpacing(20)
        
        row_label = QLabel("Row:")
        row_label.setFont(QFont("Segoe UI", 11, QFont.Bold))
        top_layout.addWidget(row_label)
        
        self.row_spin = QSpinBox()
        self.row_spin.setFont(QFont("Segoe UI", 11))
        self.row_spin.setMinimum(1)
        self.row_spin.setMaximum(1)
        self.row_spin.valueChanged.connect(self.load_row)
        top_layout.addWidget(self.row_spin)
        
        prev_btn = QPushButton("← Previous")
        prev_btn.setFont(QFont("Segoe UI", 11))
        prev_btn.clicked.connect(self.prev_row)
        top_layout.addWidget(prev_btn)
        
        next_btn = QPushButton("Next →")
        next_btn.setFont(QFont("Segoe UI", 11))
        next_btn.clicked.connect(self.next_row)
        top_layout.addWidget(next_btn)
        
        top_layout.addStretch()
        
        self.save_btn = QPushButton("💾 Save Scores")
        self.save_btn.setFont(QFont("Segoe UI", 11, QFont.Bold))
        self.save_btn.clicked.connect(self.save_scores)
        self.save_btn.setEnabled(False)
        self.save_btn.setMinimumHeight(40)
        top_layout.addWidget(self.save_btn)
        
        quit_btn = QPushButton("✕ Quit")
        quit_btn.setFont(QFont("Segoe UI", 11))
        quit_btn.clicked.connect(self.close)
        top_layout.addWidget(quit_btn)
        
        main_layout.addLayout(top_layout)
        
        # Info label
        self.info_label = QLabel("📂 Load an Excel file to begin")
        self.info_label.setFont(QFont("Segoe UI", 11))
        self.info_label.setStyleSheet("color: #0078D4; font-style: italic;")
        main_layout.addWidget(self.info_label)
        
        # Scrollable area for content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        scroll_content = QWidget()
        content_layout = QVBoxLayout(scroll_content)
        content_layout.setSpacing(15)
        
        # Source text
        eng_group = QGroupBox("📝 Source Text")
        eng_group.setFont(QFont("Segoe UI", 11, QFont.Bold))
        eng_layout = QVBoxLayout()
        
        self.english_text = QTextEdit()
        self.english_text.setFont(QFont("Segoe UI", 13))
        self.english_text.setMinimumHeight(120)
        self.english_text.setReadOnly(True)
        eng_layout.addWidget(self.english_text)
        
        eng_group.setLayout(eng_layout)
        content_layout.addWidget(eng_group)
        
        # Arabic translations - use generic labels (A, B) to hide which model is which
        display_labels = ["A", "B"]
        self.translation_groups = []  # Store group boxes to update titles if needed
        
        for i, label in enumerate(display_labels):
            group = QGroupBox(f"🌐 Translation {label}")
            group.setFont(QFont("Segoe UI", 11, QFont.Bold))
            layout = QVBoxLayout()
            
            # Arabic text display with RTL support
            text_edit = QTextEdit()
            text_edit.setFont(QFont("Arial", 15))
            text_edit.setMinimumHeight(120)
            text_edit.setReadOnly(True)
            # Enable RTL layout direction
            text_edit.setLayoutDirection(Qt.RightToLeft)
            layout.addWidget(text_edit)
            
            # Score radio buttons
            score_layout = QHBoxLayout()
            score_label = QLabel("Score:")
            score_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
            score_layout.addWidget(score_label)
            
            button_group = QButtonGroup()
            radio_buttons = []
            
            for score in range(11):
                rb = QRadioButton(str(score))
                rb.setFont(QFont("Segoe UI", 11))
                button_group.addButton(rb, score)
                radio_buttons.append(rb)
                score_layout.addWidget(rb)
            
            # Set default to 0
            radio_buttons[0].setChecked(True)
            
            score_layout.addStretch()
            layout.addLayout(score_layout)
            
            group.setLayout(layout)
            content_layout.addWidget(group)
            self.translation_groups.append(group)
            
            self.model_widgets.append({
                'display_label': label,
                'text_widget': text_edit,
                'button_group': button_group,
                'radio_buttons': radio_buttons
            })
        
        # Status label for auto-save feedback
        self.status_label = QLabel("")
        self.status_label.setFont(QFont("Segoe UI", 10))
        self.status_label.setStyleSheet("color: #107C10; font-style: italic;")
        self.status_label.setAlignment(Qt.AlignCenter)
        content_layout.addWidget(self.status_label)
        
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)
    
    def load_excel(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel file",
            "",
            "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if not filepath:
            return
        
        try:
            self.filepath = Path(filepath)
            self.wb = load_workbook(filename=str(self.filepath))
            
            # Populate sheet selector
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.wb.sheetnames)
            
            self.save_btn.setEnabled(True)
            QMessageBox.information(self, "Success", f"Loaded: {self.filepath.name}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel:\n{e}")
    
    def on_sheet_changed(self):
        sheet_name = self.sheet_combo.currentText()
        if not sheet_name or not self.wb:
            return
        
        self.current_sheet = self.wb[sheet_name]
        
        # Calculate content rows (all rows except header)
        max_row = self.current_sheet.max_row
        self.max_content_rows = max_row - 1
        
        self.row_spin.setMaximum(self.max_content_rows)
        self.row_spin.setValue(1)
        self.load_row()
    
    def load_row(self):
        if not self.current_sheet:
            return
        
        content_row_num = self.row_spin.value()
        if content_row_num < 1 or content_row_num > self.max_content_rows:
            return
        
        # Clear status label when loading new row
        self.status_label.setText("")
        
        # Excel row: header is row 1, so content starts at row 2
        excel_row_idx = 1 + content_row_num
        self.current_row_idx = excel_row_idx
        
        # Read content row
        row_data = []
        for col in range(1, 4):
            cell = self.current_sheet.cell(row=excel_row_idx, column=col)
            cell_value = cell.value
            if cell_value is None:
                row_data.append("")
            else:
                row_data.append(str(cell_value).strip())
        
        # Source text
        english = row_data[0] if row_data[0] else "(No source text)"
        self.english_text.setPlainText(english)
        
        # Store original translations (Model 1 at index 0, Model 2 at index 1)
        self.current_translations = [
            row_data[1] if len(row_data) > 1 else "",
            row_data[2] if len(row_data) > 2 else ""
        ]
        
        # Randomize display order for this row
        self.display_order = [0, 1]
        random.shuffle(self.display_order)
        
        # Find LCS matches between the two translations
        if self.current_translations[0] and self.current_translations[1]:
            matched_idx_0, matched_idx_1 = find_lcs_words(
                self.current_translations[0], 
                self.current_translations[1]
            )
        else:
            matched_idx_0, matched_idx_1 = set(), set()
        
        # Display translations in randomized order with LCS highlighting
        matched_indices_list = [matched_idx_0, matched_idx_1]
        
        for display_idx, widget_dict in enumerate(self.model_widgets):
            actual_model_idx = self.display_order[display_idx]
            translation = self.current_translations[actual_model_idx]
            matched_indices = matched_indices_list[actual_model_idx]
            
            if translation:
                self._set_text_with_highlighting(
                    widget_dict['text_widget'], 
                    translation, 
                    matched_indices
                )
            else:
                widget_dict['text_widget'].setPlainText("(No translation)")
        
        # Load existing scores from JSON if available
        self._load_scores_from_json(content_row_num - 1)
        
        # Update info
        self.info_label.setText(
            f"Sheet: {self.current_sheet.title} | Row: {content_row_num}/{self.max_content_rows} | Excel Row: {excel_row_idx}"
        )
    
    def _set_text_with_highlighting(self, text_widget, text, matched_word_indices):
        """
        Set text in widget with LCS-matched words highlighted in dark red.
        
        Args:
            text_widget: QTextEdit widget to set text in
            text: The full text string
            matched_word_indices: Set of word indices that should be highlighted
        """
        text_widget.clear()
        
        words = text.split()
        word_positions = get_word_positions(text)
        
        # Create format for highlighted (matched) words - dark red
        highlight_format = QTextCharFormat()
        highlight_format.setForeground(QColor("#8B0000"))  # Dark red
        highlight_format.setFontWeight(QFont.Bold)
        
        # Create format for normal words
        normal_format = QTextCharFormat()
        normal_format.setForeground(QColor("#000000"))  # Black
        
        cursor = text_widget.textCursor()
        
        last_pos = 0
        for word_idx, (start, end) in enumerate(word_positions):
            # Add any whitespace/text before this word
            if start > last_pos:
                cursor.insertText(text[last_pos:start], normal_format)
            
            # Add the word with appropriate formatting
            word = text[start:end]
            if word_idx in matched_word_indices:
                cursor.insertText(word, highlight_format)
            else:
                cursor.insertText(word, normal_format)
            
            last_pos = end
        
        # Add any remaining text
        if last_pos < len(text):
            cursor.insertText(text[last_pos:], normal_format)
    
    def save_scores(self, show_message=True):
        if not self.current_sheet or self.current_row_idx is None:
            if show_message:
                QMessageBox.warning(self, "No row", "No row loaded to save scores.")
            return False
        
        try:
            english_text = self.english_text.toPlainText()
            if english_text == "(No source text)":
                english_text = ""
            
            # Map display scores back to actual model order
            # display_order[i] tells us which actual model is shown at display position i
            # So we need to reverse: actual_scores[actual_model_idx] = display_score[display_idx]
            display_scores = []
            for widget_dict in self.model_widgets:
                score = widget_dict['button_group'].checkedId()
                if score == -1:
                    score = 0
                display_scores.append(score)
            
            # Convert display scores to actual model scores
            actual_scores = [0, 0]
            for display_idx, actual_model_idx in enumerate(self.display_order):
                actual_scores[actual_model_idx] = display_scores[display_idx]
            
            # Save to JSON files with actual model order
            self._save_to_json(english_text, self.current_translations, actual_scores)
            
            if show_message:
                QMessageBox.information(self, "Saved", f"Scores saved to JSON files")
            
            return True
            
        except Exception as e:
            import traceback
            error_details = f"Failed to save scores:\n{str(e)}\n\nFull traceback:\n{traceback.format_exc()}"
            print(error_details)
            if show_message:
                QMessageBox.critical(self, "Error", error_details)
            return False
    
    def _load_scores_from_json(self, row_index):
        """Load existing scores from JSON files if available, mapping to current display order."""
        sheet_name = self.current_sheet.title
        
        # First reset all to 0
        for widget_dict in self.model_widgets:
            widget_dict['button_group'].button(0).setChecked(True)
        
        if os.path.exists(DETAILED_COMMENTS_FILE):
            with open(DETAILED_COMMENTS_FILE, 'r', encoding='utf-8') as f:
                detailed_comments = json.load(f)
            
            # Build a map of model -> score for this row
            model_scores_map = {}
            for entry in detailed_comments:
                if entry.get('sheet') == sheet_name and entry.get('row_index') == row_index:
                    model = entry.get('model')
                    score = entry.get('score', 0)
                    model_scores_map[model] = score
            
            # Apply scores to widgets based on current display order
            # display_order[display_idx] = actual_model_idx
            for display_idx, actual_model_idx in enumerate(self.display_order):
                model_name = MODELS[actual_model_idx]
                if model_name in model_scores_map:
                    score = model_scores_map[model_name]
                    button = self.model_widgets[display_idx]['button_group'].button(score)
                    if button:
                        button.setChecked(True)
    
    def _save_to_json(self, english_text, translations, scores):
        """Save evaluation data to JSON files."""
        sheet_name = self.current_sheet.title
        content_row_num = self.row_spin.value()
        row_index = content_row_num - 1  # 0-based for JSON
        
        # 1. Update evaluated_rows.json
        if os.path.exists(EVALUATED_ROWS_FILE):
            with open(EVALUATED_ROWS_FILE, 'r', encoding='utf-8') as f:
                evaluated_rows = json.load(f)
        else:
            evaluated_rows = {}
        
        if sheet_name not in evaluated_rows:
            evaluated_rows[sheet_name] = []
        
        if row_index not in evaluated_rows[sheet_name]:
            evaluated_rows[sheet_name].append(row_index)
            evaluated_rows[sheet_name].sort()
        
        with open(EVALUATED_ROWS_FILE, 'w', encoding='utf-8') as f:
            json.dump(evaluated_rows, f, indent=2, ensure_ascii=False)
        
        # 2. Update model_scores.json
        if os.path.exists(MODEL_SCORES_FILE):
            with open(MODEL_SCORES_FILE, 'r', encoding='utf-8') as f:
                model_scores = json.load(f)
        else:
            model_scores = {}
        
        # Ensure all current models exist in the dictionary
        for model in MODELS:
            if model not in model_scores:
                model_scores[model] = {"total_score": 0, "count": 0, "average": 0.0, "comments_count": 0}
        
        # Check if this row was already scored (to avoid double-counting)
        # We'll track by loading detailed_comments and checking
        already_counted = False
        if os.path.exists(DETAILED_COMMENTS_FILE):
            with open(DETAILED_COMMENTS_FILE, 'r', encoding='utf-8') as f:
                detailed_comments = json.load(f)
            
            # Check if any entry matches this sheet + row_index
            for entry in detailed_comments:
                if entry.get('sheet') == sheet_name and entry.get('row_index') == row_index:
                    already_counted = True
                    # Remove old entries for this row to update them
                    detailed_comments = [e for e in detailed_comments 
                                        if not (e.get('sheet') == sheet_name and e.get('row_index') == row_index)]
                    break
        else:
            detailed_comments = []
        
        # Update model scores
        for i, (model, score) in enumerate(zip(MODELS, scores)):
            if not already_counted:
                model_scores[model]["total_score"] += score
                model_scores[model]["count"] += 1
            else:
                # Recalculate: we need to subtract old score and add new
                # Since we don't have old scores easily, we'll recalculate from detailed_comments
                pass  # Will recalculate after updating detailed_comments
            
            # Add to detailed comments
            detailed_comments.append({
                "sheet": sheet_name,
                "row_index": row_index,
                "model": model,
                "score": score,
                "comment": "",
                "english_text": english_text,
                "translation": translations[i] if i < len(translations) else ""
            })
        
        # If already counted, recalculate totals from detailed_comments
        if already_counted:
            model_scores = {}
            for model in MODELS:
                model_scores[model] = {"total_score": 0, "count": 0, "average": 0.0, "comments_count": 0}
            
            for entry in detailed_comments:
                model = entry.get('model')
                score = entry.get('score', 0)
                if model in model_scores:
                    model_scores[model]["total_score"] += score
                    model_scores[model]["count"] += 1
        
        # Calculate averages
        for model in MODELS:
            if model_scores[model]["count"] > 0:
                model_scores[model]["average"] = (
                    model_scores[model]["total_score"] / model_scores[model]["count"]
                )
        
        # Save model_scores.json
        with open(MODEL_SCORES_FILE, 'w', encoding='utf-8') as f:
            json.dump(model_scores, f, indent=2, ensure_ascii=False)
        
        # Save detailed_comments.json
        with open(DETAILED_COMMENTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(detailed_comments, f, indent=2, ensure_ascii=False)
    
    def prev_row(self):
        current = self.row_spin.value()
        if current > 1:
            self.row_spin.setValue(current - 1)
    
    def next_row(self):
        current = self.row_spin.value()
        if current < self.max_content_rows:
            # Check if both scores are non-zero for auto-save
            scores = []
            for widget_dict in self.model_widgets:
                score = widget_dict['button_group'].checkedId()
                if score == -1:
                    score = 0
                scores.append(score)
            
            # Auto-save if neither score is zero
            if all(score != 0 for score in scores):
                if self.save_scores(show_message=False):
                    self.status_label.setText("✓ Scores auto-saved")
                    self.status_label.setStyleSheet("color: #107C10; font-style: italic;")
                    # Clear the message after 2 seconds
                    QTimer.singleShot(2000, lambda: self.status_label.setText(""))
            
            self.row_spin.setValue(current + 1)
    
    def _clear_status_label(self):
        """Clear the status label text."""
        self.status_label.setText("")


def main():
    app = QApplication(sys.argv)
    window = TranslationEvaluatorQt()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
